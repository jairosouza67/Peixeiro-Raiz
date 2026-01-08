
import openpyxl
import json
import os

def dump_excel_to_json(file_path, output_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    data = {
        "file": file_path,
        "sheets": {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []
        for row in ws.iter_rows(max_row=50, max_col=20):
            row_data = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append("")
                elif isinstance(val, str) and val.startswith("="):
                    row_data.append({"formula": val})
                else:
                    try:
                        # Try to handle date objects or other non-serializable objects
                        if hasattr(val, 'isoformat'):
                             row_data.append(val.isoformat())
                        else:
                             row_data.append(val)
                    except:
                        row_data.append(str(val))
            sheet_data.append(row_data)
        data["sheets"][sheet_name] = sheet_data
        
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    dump_excel_to_json("docs/Planilha Criador Inteligente - Final.xlsx", "excel_dump.json")
