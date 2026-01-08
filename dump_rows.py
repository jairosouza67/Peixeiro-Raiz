
import openpyxl
from openpyxl.utils import get_column_letter

def dump_rows(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb['Tilápia']
    with open('detailed_rows.txt', 'w', encoding='utf-8') as f:
        for r in range(1, 40):
            row = ws[r]
            vals = []
            for c in row:
                if c.column > 30: break
                vals.append(f"{get_column_letter(c.column)}:{c.value}")
            f.write(f"Row {r}: " + " | ".join(vals) + "\n")
        
        # Also check Dashboard
        ws_dash = wb['Dashboard']
        f.write("\n--- Dashboard ---\n")
        f.write(f"D5: {ws_dash['D5'].value}\n")
        f.write(f"D6: {ws_dash['D6'].value}\n")
        f.write(f"D7: {ws_dash['D7'].value}\n")
        f.write(f"Z11: {ws_dash['Z11'].value}\n") # Wait, check if Z11 exists in Dash or Tilápia

if __name__ == "__main__":
    dump_rows("docs/Planilha Criador Inteligente - Final.xlsx")
