
import openpyxl

def extract_tilapia_table(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Tilápia']
    table = []
    # Data is in rows 4 to 36
    for r in range(4, 37):
        row = ws[r]
        row_data = {
            "stage": ws.cell(row=r, column=1).value, # A
            "startWeight": ws.cell(row=r, column=2).value, # B
            "endWeight": ws.cell(row=r, column=3).value, # C
            "pvBase": ws.cell(row=r, column=11).value, # K
            "feedType": ws.cell(row=r, column=19).value, # S
            "feedings": ws.cell(row=r, column=17).value # Q
        }
        table.append(row_data)
    
    # Temperature table AB6:AC17
    temp_table = []
    for r in range(6, 18):
        temp_table.append({
            "temp": ws.cell(row=r, column=28).value, # AB
            "factor": ws.cell(row=r, column=29).value # AC
        })
        
    import json
    with open('engine_data.json', 'w', encoding='utf-8') as f:
        json.dump({"growthTable": table, "tempTable": temp_table}, f, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    extract_tilapia_table("docs/Planilha Criador Inteligente - Final.xlsx")
