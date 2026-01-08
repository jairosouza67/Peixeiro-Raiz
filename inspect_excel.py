
import openpyxl
import os

def inspect_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    print(f"File: {file_path}")
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        # Inspect a region (e.g., 20 rows, 10 columns)
        for row in ws.iter_rows(max_row=30, max_col=15):
            row_data = []
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    row_data.append(f"F:{val}")
                else:
                    row_data.append(str(val))
            print("\t".join(row_data))

if __name__ == "__main__":
    path = "docs/Planilha Criador Inteligente - Final.xlsx"
    if os.path.exists(path):
        inspect_excel(path)
    else:
        print(f"File not found: {path}")
